#!/usr/bin/env python3
"""把「gal 合集主」xlsx 导入 mo 站游戏区的 Galgame 小分区。

和 import_from_xlsx.py 完全独立，互不覆盖：
  - 那个脚本管 id 带 `xlsx-` 前缀的条目，重跑会清掉所有 `xlsx-*`
  - 这个脚本管 id 带 `gal-` 前缀的条目，重跑只清 `gal-*`
两边各扫各的，不会互相清空。

**必须用 openpyxl 读超链接。** 这张表的网盘地址是 Excel 超链接对象，
单元格文字只是「夸克网盘分享」这类说明词，正则扫文本一条都扫不出来
（import_from_xlsx.py 就是扫文本的，对这张表无效）。

用法：
    python import_gal_xlsx.py "D:\\game\\...\\gal合集主 - 录入.xlsx" --dry-run
    python import_gal_xlsx.py "D:\\game\\...\\gal合集主 - 录入.xlsx"

id 由「名字的哈希」生成而不是行号 —— 表格里插删行、重新排序时 id 不变，
点击统计不会串到别的作品上。改名字会被当成「删一条 + 加一条」。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("需要 openpyxl：pip install openpyxl")

IMPORT_PREFIX = "gal-"
HERE = Path(__file__).resolve().parent
DST = HERE / "data" / "items.json"

# 三块内容横向并排，不是一张连续的表。行号取自实际表格结构。
BLOCK_MAIN = (3, 360)     # A=名字 B=链接 C=解压码
BLOCK_NEW = (9, 41)       # D=名字 E=链接 F=解压码
BLOCK_CLUB = (366, 535)   # A=编号 B=名字 G=平台 H=链接

PAN_NAMES = {
    "pan.baidu.com": "百度网盘",
    "drive.uc.cn": "UC 网盘",
    "pan.quark.cn": "夸克网盘",
    "pan.xunlei.com": "迅雷网盘",
    "yun.139.com": "移动云盘",
    "www.kdocs.cn": "金山文档",
    "kdocs.cn": "金山文档",
}

# 平台标记 -> 展示名。按长度降序匹配，"安卓直装" 要先于 "安卓"。
PLATFORM_TOKENS = [
    ("安卓使用电脑模拟器", "安卓模拟器"),
    ("安卓直装", "安卓直装"),
    ("模拟器", "模拟器"),
    ("switch", "Switch"),
    ("linux", "Linux"),
    ("安卓", "安卓"),
    ("苹果", "苹果"),
    ("直装", "直装"),
    ("psv", "PSV"),
    ("ios", "iOS"),
    ("ips", "iOS"),  # 表里把 ios 打成了 ips
    ("apk", "安卓"),
    ("mod", "MOD"),
    ("双端", "双端"),
    ("pc", "PC"),
    ("az", "安卓"),
    ("kr", "kr 模拟器"),
    ("ty", "ty 模拟器"),
    ("ons", "ons 模拟器"),
]

# 这几行是解压工具/模拟器，不是游戏，单独归到工具区且不标成人向。
TOOL_KEYWORDS = ["解压教程", "解压软件", "手机模拟器"]

# 表头、群公告、栏目标题之类的非资源行。整体匹配名字，不做子串误伤。
NOT_A_RESOURCE = [
    "上千galgame大合集", "资源来源网络，如有侵权联系删除", "游戏文件⬇",
    "失效了去群里找群主补！", "【Galgame合集】-宝子们请自行搜索",
    "热门GAL,每日更新且可以每天保存一次，防止资源失效", "杂gal",
]

SEP = r"[\s/+、\\，,]*"


def pan_name(url: str) -> str:
    for domain, label in PAN_NAMES.items():
        if domain in url:
            return label
    return "网盘"


def norm_space(text: str) -> str:
    """表里混着 nbsp(\\xa0)、全角空格和连续空格，统一压成单个半角空格。
    不压的话名字里会留下 '\\xa0 魔女的夜宴' 这种前导空白。"""
    return re.sub(r"[\s　\xa0]+", " ", (text or "")).strip()


def split_platforms(raw: str) -> tuple[str, list[str]]:
    """把「pc/安卓/kr  某作品名」拆成 (作品名, 平台标签)。

    平台前缀只在标题**开头**识别 —— 作品名里也可能出现 "pc" 这类字母
    （比如某些西文标题），从中间乱切会把名字截断。
    """
    text = norm_space(raw)
    tags: list[str] = []
    # 逐个吃掉开头的平台词，直到遇到非平台内容
    changed = True
    while changed and text:
        changed = False
        head = text.lstrip(" /\\、,，+")
        if head != text:
            text = head
            changed = True
        low = text.lower()
        for token, label in PLATFORM_TOKENS:
            if low.startswith(token):
                # "pc" 不能吃掉 "pcゲーム" 这种粘连的西文词
                nxt = low[len(token):len(token) + 1]
                if nxt and nxt.isalpha() and token.isalpha() and nxt not in "/\\":
                    continue
                if label not in tags:
                    tags.append(label)
                text = text[len(token):]
                changed = True
                break
    # 只在开头去连接符；结尾的 "-" 可能是标题的一部分
    # （「夏日口袋 -REFLECTION BLUE-」那个尾巴不能丢）
    name = norm_space(text).lstrip(" /\\、,，+-—·").rstrip(" /\\、,，+·")
    return (name or norm_space(raw)), tags


def platforms_from_cell(raw: str) -> list[str]:
    """块3 的平台单独在 G 列，形如 "PC/安卓/PSV/IOS"。"""
    tags: list[str] = []
    for part in re.split(r"[\s/、,，+]+", raw or ""):
        low = part.strip().lower()
        if not low:
            continue
        for token, label in PLATFORM_TOKENS:
            if low == token:
                if label not in tags:
                    tags.append(label)
                break
    return tags


def clean_password(raw: str) -> str:
    """C/F 列形如「解压码：0721」，取出码本身。

    注意这是**压缩包解压码**，不是网盘提取码 —— 站点 `password` 字段在卡片上
    显示为「提取码」，把解压码放进去会误导访客。所以解压码走 note，
    `password` 只放网盘链接自带的 pwd。
    """
    m = re.search(r"([A-Za-z0-9]{4,10})\s*$", (raw or "").strip())
    return m.group(1) if m else ""


def share_pwd(url: str) -> str:
    """网盘链接里自带的提取码，形如 ...?pwd=7h44。"""
    m = re.search(r"[?&]pwd=([A-Za-z0-9]+)", url or "")
    return m.group(1) if m else ""


def make_id(name: str) -> str:
    """按名字生成稳定 id。行号会因插删行而变，名字不会。"""
    h = hashlib.sha256(name.encode("utf-8")).hexdigest()[:10]
    return f"{IMPORT_PREFIX}{h}"


def is_tool(name: str) -> bool:
    return any(k in name for k in TOOL_KEYWORDS)


def is_club_bundle(name: str) -> bool:
    """块1 里也散着「柚子社 合集」这类整社打包，和块3 的社团合集同性质。"""
    return bool(re.fullmatch(r"[\w一-鿿]{1,12}社\s*合集", name))


def make_item(
    name: str, url: str, unzip_pw: str, platforms: list[str],
    kind: str, note: str = "", extra_tags: list[str] | None = None,
) -> dict:
    """一条资源。adult 默认 True —— 这个合集整体是成人向 galgame，
    漏标比多标严重（多标只是要切成年模式才看见，漏标是直接暴露给未成年）。
    工具类走 is_tool() 分流，不到这里。

    `password` 只放网盘自带提取码；解压码写进 note，因为卡片上那个字段
    显示的是「提取码」，塞解压码会让人拿去填网盘密码框。
    """
    tags = list(platforms)
    for t in extra_tags or []:
        if t not in tags:
            tags.append(t)
    if url:
        pn = pan_name(url)
        if pn not in tags:
            tags.append(pn)

    notes = [note] if note else []
    if unzip_pw:
        notes.append(f"压缩包解压码：{unzip_pw}（不是网盘提取码）。")

    item = {
        "id": make_id(name),
        "name": name,
        "section": "game",
        "subsection": "gal",
        "description": (
            f"{'、'.join(platforms)} 平台的 Galgame 资源。" if platforms
            else "Galgame 资源。"
        ),
        "url": url,
        "tags": tags[:6],
        "kind": kind,
        "need_login": False,
        "update_info": "网盘分享",
        "note": " ".join(notes),
        "adult": True,
    }
    pw = share_pwd(url)
    if pw:
        item["password"] = pw
    return item


def make_tool(name: str, url: str, unzip_pw: str, plats: list[str]) -> dict:
    """解压工具/模拟器。不标 adult —— 工具本身无内容。"""
    tags = list(plats)
    if url:
        pn = pan_name(url)
        if pn not in tags:
            tags.append(pn)
    item = {
        "id": make_id(name),
        "name": name,
        "section": "tool",
        "description": "Galgame 合集配套的解压 / 模拟器工具。",
        "url": url,
        "tags": tags[:6],
        "kind": "工具 / 网站",
        "need_login": False,
        "update_info": "网盘分享",
        "note": f"压缩包解压码：{unzip_pw}（不是网盘提取码）。" if unzip_pw else "",
    }
    pw = share_pwd(url)
    if pw:
        item["password"] = pw
    return item


def cell_link(ws, row: int, col: int) -> str:
    """取超链接目标。单元格文字是「夸克网盘分享」这类说明词，不含地址。"""
    cell = ws.cell(row, col)
    return cell.hyperlink.target.strip() if cell.hyperlink else ""


def cell_text(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    return norm_space(str(v)) if v is not None else ""


def parse_main_block(ws) -> list[dict]:
    """块1：A=名字（含平台前缀） B=网盘链接 C=解压码。表里最大的一块。"""
    items = []
    for r in range(BLOCK_MAIN[0], BLOCK_MAIN[1] + 1):
        raw = cell_text(ws, r, 1)
        if not raw:
            continue
        url = cell_link(ws, r, 2)
        pw = clean_password(cell_text(ws, r, 3))
        name, plats = split_platforms(raw)

        if name in NOT_A_RESOURCE or name.startswith("http"):
            continue
        # 表头说明、群公告之类的行：没链接又不像作品名
        if not url and (len(name) > 40 or any(
            k in name for k in ["请自行搜索", "失效了去群里", "如有侵权", "为什么有的没链接",
                                "游戏名字后面", "每日更新且可以"]
        )):
            continue

        if is_tool(name):
            items.append(make_tool(name, url, pw, plats))
            continue

        # galN / gal补充N 是打包卷，不是单部作品
        if re.fullmatch(r"gal\s*补充\s*\d+|gal\s*\d+\s*g?a?\s*l?", name, re.I):
            items.append(make_item(
                name, url, pw, plats, "Galgame 打包卷",
                note="整卷打包，内含多部作品。",
                extra_tags=["打包卷"],
            ))
            continue

        # 「柚子社 合集」这类整社打包
        if is_club_bundle(name):
            items.append(make_item(
                name, url, pw, plats, "社团合集",
                note="整社打包，内含多部作品。",
                extra_tags=["社团合集"],
            ))
            continue

        note = "" if url else "表格里这条没给链接，等表格补上后重跑导入即可。"
        items.append(make_item(name, url, pw, plats, "Galgame 资源", note=note))
    return items


def parse_new_block(ws) -> list[dict]:
    """块2：D=名字 E=链接 F=解压码。多数只有名字没链接，按用户要求保留。"""
    items = []
    for r in range(BLOCK_NEW[0], BLOCK_NEW[1] + 1):
        raw = cell_text(ws, r, 4)
        if not raw:
            continue
        url = cell_link(ws, r, 5)
        pw = clean_password(cell_text(ws, r, 6))
        name, plats = split_platforms(raw)
        if not name or name.startswith("http"):
            continue
        # 说明行
        if any(k in name for k in ["单击网盘分享", "如有侵权", "打开文件教程"]):
            continue
        if is_tool(name) or "解压软件" in name:
            items.append(make_tool(name, url, pw, plats))
            continue
        note = "" if url else "表格里这条没给链接，等表格补上后重跑导入即可。"
        items.append(make_item(
            name, url, pw, plats, "Galgame 资源", note=note, extra_tags=["新增区"]))
    return items


def parse_club_block(ws) -> list[dict]:
    """块3：A=编号 B=名字 G=平台 H=链接。按社团分组，组头是「XX社合集」。"""
    items = []
    club = ""
    for r in range(BLOCK_CLUB[0], BLOCK_CLUB[1] + 1):
        code = cell_text(ws, r, 1)
        name = cell_text(ws, r, 2)
        if not name:
            continue
        url = cell_link(ws, r, 8)
        plats = platforms_from_cell(cell_text(ws, r, 7))

        # 组头：编号形如 "key社/a01"，不是 "a01001" 这种作品号
        if code and not re.fullmatch(r"[a-z]\d{3,}", code, re.I):
            club = code.split("/")[0].strip()
            items.append(make_item(
                name, url, "", plats, "社团合集",
                note=f"{club}整社打包，内含多部作品。",
                extra_tags=["社团合集", club],
            ))
            continue

        note = "" if url else "表格里这条没给链接，等表格补上后重跑导入即可。"
        items.append(make_item(
            name, url, "", plats, "Galgame 资源",
            note=note, extra_tags=[club] if club else None))
    return items


def merge_same_name(items: list[dict]) -> list[dict]:
    """同名条目合并成多网盘源。表里同一作品可能在不同块各给一个网盘。"""
    out: dict[str, dict] = {}
    for it in items:
        key = it["id"]
        if key not in out:
            it = dict(it)
            if it.get("url"):
                it["links"] = [{
                    "name": pan_name(it["url"]),
                    "url": it["url"],
                    "password": it.get("password", ""),
                }]
            out[key] = it
            continue
        # 已有同名：把新的网盘源并进去
        cur = out[key]
        if not it.get("url"):
            continue
        links = cur.setdefault("links", [])
        if any(l["url"] == it["url"] for l in links):
            continue
        links.append({
            "name": pan_name(it["url"]),
            "url": it["url"],
            "password": it.get("password", ""),
        })
        if not cur.get("url"):
            cur["url"] = it["url"]
        for t in it.get("tags", []):
            if t not in cur["tags"] and len(cur["tags"]) < 6:
                cur["tags"].append(t)

    # 单源的不需要 links，前端会用 url 自动兜底
    for it in out.values():
        links = it.get("links") or []
        if len(links) <= 1:
            it.pop("links", None)
            continue
        # 同一网盘出现多次（表里同作品给了两个迅雷链接），加序号区分，
        # 否则详情页会出现两个一模一样的「迅雷网盘」按钮，分不清点哪个
        seen: Counter = Counter(l["name"] for l in links)
        used: Counter = Counter()
        for l in links:
            if seen[l["name"]] > 1:
                used[l["name"]] += 1
                l["label"] = f"源 {used[l['name']]}"
    return list(out.values())


def load_dst() -> dict:
    if not DST.exists():
        sys.exit(f"找不到 {DST}")
    return json.loads(DST.read_text(encoding="utf-8"))


def main() -> int:
    ap = argparse.ArgumentParser(description="把 gal 合集 xlsx 导入 mo 站游戏区")
    ap.add_argument("xlsx", help="xlsx 文件路径")
    ap.add_argument("--dry-run", action="store_true", help="只预览，不写文件")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        sys.exit(f"找不到文件：{xlsx}")

    wb = openpyxl.load_workbook(xlsx)
    ws = wb.worksheets[0]
    print(f"工作表：{ws.title}（{ws.max_row} 行）")

    raw_items = parse_main_block(ws) + parse_new_block(ws) + parse_club_block(ws)
    print(f"三块共解析出 {len(raw_items)} 条")

    new_items = merge_same_name(raw_items)
    merged = len(raw_items) - len(new_items)
    if merged:
        print(f"同名合并 {merged} 条（多网盘源并入 links）")

    if not new_items:
        sys.exit("解析结果为空，中止 —— 避免把已有条目清空")

    no_link = [it for it in new_items if not it.get("url")]
    dist = Counter((it["section"], it.get("subsection") or "-") for it in new_items)
    kinds = Counter(it["kind"] for it in new_items)
    print("\n分区分布：")
    for (sec, sub), n in sorted(dist.items()):
        print(f"  {sec}/{sub}: {n}")
    print("类型分布：")
    for k, n in kinds.most_common():
        print(f"  {k}: {n}")
    print(f"成人向标记：{sum(1 for it in new_items if it.get('adult'))} / {len(new_items)}")
    print(f"无链接条目：{len(no_link)}（保留，note 里已说明）")

    dst = load_dst()
    kept = [
        it for it in (dst.get("items") or [])
        if not str(it.get("id", "")).startswith(IMPORT_PREFIX)
    ]
    old = {
        it["id"]: it
        for it in (dst.get("items") or [])
        if str(it.get("id", "")).startswith(IMPORT_PREFIX)
    }
    new_ids = {it["id"] for it in new_items}
    added = [it for it in new_items if it["id"] not in old]
    removed = [it for it in old.values() if it["id"] not in new_ids]

    print(f"\n对比上次导入：新增 {len(added)}，移除 {len(removed)}，"
          f"其余 {len(new_items) - len(added)} 条按最新表格刷新")
    print(f"非 {IMPORT_PREFIX} 条目 {len(kept)} 条保留不动")
    for it in added[:8]:
        print(f"  + {it['name'][:46]}")
    for it in removed[:8]:
        print(f"  - {str(it.get('name'))[:46]}")

    if args.dry_run:
        print("\n--dry-run：未写入文件")
        return 0

    dst["items"] = kept + new_items
    dst["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    tmp = DST.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(dst, fh, ensure_ascii=False, indent=2)
        fh.write("\n")
    tmp.replace(DST)

    print(f"\n已写入 {DST.name}，共 {len(dst['items'])} 条")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
